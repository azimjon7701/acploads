import openpyxl

path = "utils/external_files/uscities.xlsx"
def load_excel():
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active

    cell_obj = sheet_obj.cell(row=1, column=1)

    places: list = []
    for i in range(2,30411):
        city_name = sheet_obj.cell(row=i, column=1).value
        state_code = sheet_obj.cell(row=i, column=3).value
        latitude = sheet_obj.cell(row=i, column=7).value
        longitude = sheet_obj.cell(row=i, column=8).value
        if city_name and state_code and latitude and longitude:
            name = f"{city_name}, {state_code}, USA"
            places.append((name, latitude, longitude))
    return places


# print(load_excel())